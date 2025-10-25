import os
import time
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

import webClick


# ---------------------- 1. 日志配置（自动创建日志文件，记录所有输出） ----------------------
def setup_logger():
    """配置日志：同时输出到控制台和日志文件（booking_logs.txt）"""
    # 日志文件路径（与程序同目录）
    log_file = "booking_logs.log"

    # 创建日志器
    logger = logging.getLogger("BookingChecker")
    logger.setLevel(logging.INFO)  # 日志级别：INFO及以上（包括INFO、WARNING、ERROR等）

    # 避免重复添加处理器
    if logger.hasHandlers():
        logger.handlers.clear()

    # 格式：时间 + 日志级别 + 消息（如：2025-10-24 15:30:00 [INFO] 程序启动）
    formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    # 控制台处理器（输出到屏幕）
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)

    # 文件处理器（输出到日志文件，追加模式）
    file_handler = logging.FileHandler(log_file, mode='a', encoding='utf-8')
    file_handler.setFormatter(formatter)

    # 添加处理器
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    return logger


# 初始化日志器（全局使用）
logger = setup_logger()


# ---------------------- 2. Excel处理核心函数（使用logger输出日志） ----------------------
def init_excel_status(excel_path):
    REQUIRED_HEADERS = [
        "预约开始日期", "预约结束日期", "每天时间开始",
        "每天时间结束", "是否开启检测预约", "状态"
    ]
    if not os.path.exists(excel_path):
        logger.error(f"Excel文件不存在：{excel_path}")
        return False, {}
    try:
        workbook = load_workbook(excel_path, read_only=False, data_only=False)
        worksheet = workbook.active
        header_col_map = {}
        missing_headers = []
        header_row = 1
        for col in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            header_text = str(worksheet[f"{col_letter}{header_row}"].value).strip() if worksheet[
                f"{col_letter}{header_row}"].value else ""
            if header_text in REQUIRED_HEADERS:
                header_col_map[header_text] = col
        for header in REQUIRED_HEADERS:
            if header not in header_col_map:
                missing_headers.append(header)
        if missing_headers:
            logger.error(f"Excel缺少必填表头：{', '.join(missing_headers)}")
            workbook.close()
            return False, {}
        status_col = header_col_map["状态"]
        status_col_letter = get_column_letter(status_col)
        logger.info(f"Excel表头校验通过，状态列位于第{status_col}列（{status_col_letter}列）")
        # 新增：将“已有其他时间完成”加入合法状态列表
        valid_statuses = ["待执行", "执行中", "执行成功", "执行失败", "配置无效", "已有其他时间完成"]
        for row in range(header_row + 1, worksheet.max_row + 1):
            current_status = worksheet[f"{status_col_letter}{row}"].value
            # 仅重置“非合法状态”的行
            if not current_status or current_status not in valid_statuses:
                worksheet[f"{status_col_letter}{row}"].value = "待执行"
                worksheet[f"{status_col_letter}{row}"].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6",
                                                                          fill_type="solid")
                worksheet[f"{status_col_letter}{row}"].font = Font(color="000000")
        workbook.save(excel_path)
        workbook.close()
        logger.info("Excel状态列初始化完成，所有未处理行已设为「待执行」")
        return True, header_col_map
    except Exception as e:
        logger.error(f"Excel初始化失败：{str(e)}")
        return False, {}

def validate_excel_row(excel_path, header_col_map, row):
    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=False)
        worksheet = workbook.active
        row_config = {}
        is_valid = True

        start_date_col = get_column_letter(header_col_map["预约开始日期"])
        start_date_raw = str(worksheet[f"{start_date_col}{row}"].value).strip() if worksheet[
            f"{start_date_col}{row}"].value else ""
        end_date_col = get_column_letter(header_col_map["预约结束日期"])
        end_date_raw = str(worksheet[f"{end_date_col}{row}"].value).strip() if worksheet[
            f"{end_date_col}{row}"].value else ""
        daily_start_col = get_column_letter(header_col_map["每天时间开始"])
        daily_start_raw = str(worksheet[f"{daily_start_col}{row}"].value).strip() if worksheet[
            f"{daily_start_col}{row}"].value else ""
        daily_end_col = get_column_letter(header_col_map["每天时间结束"])
        daily_end_raw = str(worksheet[f"{daily_end_col}{row}"].value).strip() if worksheet[
            f"{daily_end_col}{row}"].value else ""
        enable_col = get_column_letter(header_col_map["是否开启检测预约"])
        enable_str = str(worksheet[f"{enable_col}{row}"].value).strip() if worksheet[f"{enable_col}{row}"].value else ""

        if enable_str not in ["是", "否", "已完成"]:
            logger.error(f"第{row}行无效：「是否开启检测预约」必须为'是'、'否'或'已完成'（当前：{enable_str}）")
            is_valid = False
        row_config["enable_check"] = enable_str

        if enable_str == "是":
            try:
                start_date_clean = start_date_raw[:10]
                row_config["start_date"] = datetime.strptime(start_date_clean, "%Y-%m-%d").date()
            except ValueError:
                logger.error(f"第{row}行无效：「预约开始日期」格式错误（需类似2025-11-25，当前：{start_date_raw}）")
                is_valid = False

            try:
                end_date_clean = end_date_raw[:10]
                row_config["end_date"] = datetime.strptime(end_date_clean, "%Y-%m-%d").date()
            except ValueError:
                logger.error(f"第{row}行无效：「预约结束日期」格式错误（需类似2025-11-25，当前：{end_date_raw}）")
                is_valid = False

            if is_valid and row_config["start_date"] > row_config["end_date"]:
                logger.error(f"第{row}行无效：「预约开始日期」不能晚于「预约结束日期」")
                is_valid = False

            try:
                daily_start_clean = daily_start_raw[:5]
                row_config["daily_start_time"] = datetime.strptime(daily_start_clean, "%H:%M").time()
            except ValueError:
                logger.error(f"第{row}行无效：「每天时间开始」格式错误（需类似08:30，当前：{daily_start_raw}）")
                is_valid = False

            try:
                daily_end_clean = daily_end_raw[:5]
                row_config["daily_end_time"] = datetime.strptime(daily_end_clean, "%H:%M").time()
            except ValueError:
                logger.error(f"第{row}行无效：「每天时间结束」格式错误（需类似15:30，当前：{daily_end_raw}）")
                is_valid = False

            if is_valid and row_config["daily_start_time"] > row_config["daily_end_time"]:
                logger.error(f"第{row}行无效：「每天时间开始」不能晚于「每天时间结束」")
                is_valid = False

        workbook.close()
        return is_valid, row_config if is_valid else {}

    except Exception as e:
        logger.error(f"第{row}行校验异常：{str(e)}")
        return False, {}


def update_excel_status(excel_path, header_col_map, row, status, update_enable_col=False):
    # 新增：将“已有其他时间完成”加入允许的状态列表
    ALLOWED_STATUSES = ["执行中", "执行成功", "执行失败", "配置无效", "已有其他时间完成"]
    if status not in ALLOWED_STATUSES:
        logger.error(f"无效状态值：{status}")
        return False
    try:
        workbook = load_workbook(excel_path, read_only=False, data_only=False)
        worksheet = workbook.active
        status_col = header_col_map["状态"]
        status_col_letter = get_column_letter(status_col)
        cell = worksheet[f"{status_col_letter}{row}"]
        cell.value = status
        # 新增：“已有其他时间完成”的样式（示例：橙色填充+深橙字体，可自定义）
        status_styles = {
            "执行中": {"fill": "FFFFCC", "font": "FF6600"},
            "执行成功": {"fill": "E6F3FF", "font": "0066CC"},
            "执行失败": {"fill": "FFE6E6", "font": "CC0000"},
            "配置无效": {"fill": "F2F2F2", "font": "666666"},
            "已有其他时间完成": {"fill": "FFF2CC", "font": "FF9900"}  # 新状态样式
        }
        style = status_styles[status]
        cell.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
        cell.font = Font(color=style["font"])
        # 原有逻辑：更新“是否开启检测预约”列（仅执行成功时）
        if update_enable_col:
            enable_col = get_column_letter(header_col_map["是否开启检测预约"])
            worksheet[f"{enable_col}{row}"].value = "已完成"
            logger.info(f"第{row}行「是否开启检测预约」已更新为：「已完成」")
        workbook.save(excel_path)
        workbook.close()
        logger.info(f"第{row}行状态已更新为：「{status}」")
        return True
    except Exception as e:
        logger.error(f"第{row}行状态更新失败：{str(e)}")
        return False

# ---------------------- 3. 单次检测任务（被循环调用） ----------------------
def run_single_check(excel_path, header_col_map):
    """执行一次完整的检测任务（遍历所有行）"""
    logger.info("\n" + "=" * 50)
    logger.info("开始新一轮检测任务")
    logger.info("=" * 50)
    # 新增：存储已成功完成预约的驾照编号（单次检测内生效）
    completed_licence_nums = []
    try:
        workbook = load_workbook(excel_path, read_only=False, data_only=False)
        worksheet = workbook.active
        header_row = 1
        total_rows = worksheet.max_row
        # 遍历所有数据行
        for row in range(header_row + 1, total_rows + 1):
            logger.info(f"\n--- 处理第{row}行任务 ---")
            # 新增：1. 读取当前行驾照编号，先校验是否已存在成功记录
            current_dl_num = str(worksheet[f"A{row}"].value).strip() if worksheet[f"A{row}"].value else ""
            if current_dl_num in completed_licence_nums:
                logger.info(f"第{row}行驾照编号[{current_dl_num}]已存在成功预约，标记为「已有其他时间完成」")
                update_excel_status(excel_path, header_col_map, row, "已有其他时间完成")
                continue  # 跳过后续处理

            # 原有逻辑：2. 校验当前行配置
            row_valid, row_config = validate_excel_row(excel_path, header_col_map, row)
            if not row_valid:
                update_excel_status(excel_path, header_col_map, row, "配置无效")
                continue

            # 原有逻辑：3. 跳过已完成或未开启的任务
            if row_config["enable_check"] in ["已完成", "否"]:
                # 补充：若已完成，将驾照编号加入已完成列表（避免漏判）
                if row_config["enable_check"] == "已完成" and current_dl_num:
                    completed_licence_nums.append(current_dl_num)
                logger.info(f"第{row}行状态为「{row_config['enable_check']}」，跳过处理")
                continue

            # 原有逻辑：4. 标记为执行中
            update_success = update_excel_status(excel_path, header_col_map, row, "执行中")
            if not update_success:
                logger.warning(f"第{row}行状态更新失败，跳过")
                continue

            # 原有逻辑：5. 读取当前行配置数据
            try:
                config_data = {
                    "dlNumber": current_dl_num,  # 复用已读取的驾照编号
                    "contactName": str(worksheet[f"B{row}"].value).strip() if worksheet[f"B{row}"].value else "",
                    "contactPhone": str(worksheet[f"C{row}"].value).strip() if worksheet[f"C{row}"].value else "",
                    "Test type": str(worksheet[f"D{row}"].value).strip() if worksheet[f"D{row}"].value else "",
                    "Region": str(worksheet[f"E{row}"].value).strip() if worksheet[f"E{row}"].value else "",
                    "Centre": str(worksheet[f"F{row}"].value).strip() if worksheet[f"F{row}"].value else "",
                    "contactEmail": str(worksheet[f"G{row}"].value).strip() if worksheet[f"G{row}"].value else "",
                    "CardNumber": str(worksheet[f"H{row}"].value).strip() if worksheet[f"H{row}"].value else "",
                    "ExpiryMonth": str(worksheet[f"I{row}"].value).strip() if worksheet[f"I{row}"].value else "",
                    "ExpiryYear": str(worksheet[f"J{row}"].value).strip() if worksheet[f"J{row}"].value else "",
                    "CVN": str(worksheet[f"K{row}"].value).strip() if worksheet[f"K{row}"].value else ""
                }
            except Exception as e:
                logger.error(f"读取第{row}行配置数据失败：{str(e)}")
                update_excel_status(excel_path, header_col_map, row, "执行失败")
                continue

            # 原有逻辑：6. 执行检测任务
            task_success = False
            try:
                logger.info(
                    f"执行检测（日期：{row_config['start_date']}~{row_config['end_date']}，时间：{row_config['daily_start_time']}~{row_config['daily_end_time']}）")
                task_success = webClick.openweb(
                    start_date=row_config['start_date'],
                    end_date=row_config['end_date'],
                    daily_start_time=row_config['daily_start_time'],
                    daily_end_time=row_config['daily_end_time'],
                    config_data=config_data
                )
            except Exception as e:
                logger.error(f"检测任务异常：{str(e)}")
                task_success = False

            # 原有逻辑：7. 更新最终状态
            final_status = "执行成功" if task_success else "执行失败"
            update_result = update_excel_status(
                excel_path,
                header_col_map,
                row,
                final_status,
                update_enable_col=(final_status == "执行成功")
            )

            # 新增：8. 若执行成功，处理同驾照其他任务
            if task_success and update_result and current_dl_num:
                # 将当前驾照编号加入已完成列表
                completed_licence_nums.append(current_dl_num)
                logger.info(f"\n--- 开始取消同驾照[{current_dl_num}]的其他任务 ---")
                # 重新遍历所有行，更新同驾照的“待执行/执行中”任务
                for other_row in range(header_row + 1, total_rows + 1):
                    if other_row == row:
                        continue  # 跳过当前成功行
                    # 读取其他行的驾照编号
                    other_dl_num = str(worksheet[f"A{other_row}"].value).strip() if worksheet[
                        f"A{other_row}"].value else ""
                    # 读取其他行的当前状态
                    other_status_col_letter = get_column_letter(header_col_map["状态"])
                    other_current_status = worksheet[f"{other_status_col_letter}{other_row}"].value
                    # 条件：驾照编号相同 + 状态为“待执行”或“执行中”
                    if other_dl_num == current_dl_num and other_current_status in ["待执行", "执行中"]:
                        logger.info(f"第{other_row}行同驾照[{current_dl_num}]，更新为「已有其他时间完成」")
                        # 更新状态（无需修改“是否开启检测预约”列）
                        update_excel_status(
                            excel_path,
                            header_col_map,
                            other_row,
                            "已有其他时间完成",
                            update_enable_col=False  # 仅标记状态，不修改开启状态
                        )
                logger.info(f"--- 同驾照[{current_dl_num}]其他任务取消完成 ---")

        workbook.close()
        logger.info(f"本轮检测任务完成")
    except Exception as e:
        logger.error(f"单次检测任务异常：{str(e)}")
        try:
            workbook.close()
        except:
            pass
# ---------------------- 4. 主循环（每分钟执行一次） ----------------------
def main():
    # 配置Excel路径和循环间隔
    EXCEL_PATH = "./预约配置表.xlsx"  # 替换为你的Excel路径
    CHECK_INTERVAL = 60  # 循环间隔（秒）

    # 初始化Excel
    logger.info("=== 初始化程序 ===")
    excel_valid, header_col_map = init_excel_status(EXCEL_PATH)
    if not excel_valid:
        logger.error("初始化失败，程序退出")
        return

    # 开始循环检测
    logger.info(f"\n=== 开始循环检测（每{CHECK_INTERVAL}秒一次）===")
    logger.info("=== 按 Ctrl+C 可停止程序 ===")
    try:
        while True:
            # 执行一次检测任务
            run_single_check(EXCEL_PATH, header_col_map)

            # 等待指定时间（下次检测前）
            logger.info(f"\n等待{CHECK_INTERVAL}秒后进行下一轮检测...\n")
            time.sleep(CHECK_INTERVAL)

    except KeyboardInterrupt:
        # 捕获Ctrl+C，优雅退出
        logger.info("\n=== 用户中断程序 ===")
        logger.info("程序已停止运行")


if __name__ == "__main__":
    main()